"""The Product column: drapery or blinds, read from HD's vendor item numbers.

The shapes asserted here are real numbers taken off HD's 850s, not invented
ones -- 138VB48D36WHTC is a 1 3/8" vinyl blind, 72318-109-52-84-404 a drapery
style. A regex that passes on invented input would not tell us much.
"""
from openpyxl import load_workbook

import io

from app.products import BLIND, DRAPE, MIXED, label_for, product_of, vendor_items_in
from app.report import build_workbook, extract, product_for


def test_blind_item_numbers():
    assert product_of("138VB48D36WHTC") == BLIND      # 1 3/8" vinyl blind
    assert product_of("138VB7272WHTC") == BLIND       # no fraction letter in the size
    assert product_of("020FW12X48WHT") == BLIND       # faux wood
    assert product_of("138vb48d36whtc") == BLIND      # case is not guaranteed


def test_drape_style_numbers_of_every_block_count():
    assert product_of("72136-109-001") == DRAPE                # three blocks
    assert product_of("71555-109-644-108") == DRAPE            # four
    assert product_of("72318-109-52-84-404") == DRAPE          # five
    assert product_of("70915-170-001-108") == DRAPE            # a 70xxx style


def test_unrecognised_numbers_are_not_guessed_at():
    """A number that fits neither shape must not be filed under the line it
    resembles -- 069556586000 is a UPC, and 8xxxx would be a style block we
    have never been sent."""
    assert product_of("069556586000") == ""
    assert product_of("83001-109-001") == ""
    assert product_of("") == ""
    assert product_of(None) == ""


def test_label_reports_mixed_only_when_both_lines_are_present():
    assert label_for(["72136-109-001", "71555-109-644-108"]) == DRAPE
    assert label_for(["138VB48D36WHTC", "138VB7272WHTC"]) == BLIND
    assert label_for(["138VB48D36WHTC", "72136-109-001"]) == MIXED


def test_one_unplaceable_item_does_not_make_a_po_mixed():
    """Calling a PO of blinds Mixed would send accounting looking for drapery
    that is not on it."""
    assert label_for(["138VB48D36WHTC", "SOMETHINGNEW"]) == BLIND
    assert label_for(["SOMETHINGNEW"]) == ""


def test_vendor_items_read_from_an_850():
    detail = {"file": {"generic_json_edi": {"detail": {"baseline_item_data_loop": [
        {"baseline_item_data": {"line_item_number": 10, "vendors_item_number": "138VB48D36WHTC"}},
        {"baseline_item_data": {"line_item_number": 20, "vendors_item_number": "72136-109-001"}},
        {"baseline_item_data": {"line_item_number": 30}},   # DSD lines carry none
    ]}}}}
    assert vendor_items_in(detail) == ["138VB48D36WHTC", "72136-109-001"]
    assert vendor_items_in({}) == []


def test_dsd_is_drapery_without_an_850_to_read():
    """DSD 850s carry no vendors_item_number at all -- Crstl's CSV export has
    the field, generic_json_edi does not -- and every DSD item HD has ordered
    is drapery."""
    assert product_for("DSD", {}) == DRAPE
    assert product_for("DSD", None) == DRAPE


def test_dropship_without_a_po_on_file_is_unknown_not_assumed():
    assert product_for("Dropship", {}) == "Unknown"
    assert product_for("Wholesale", None) == "Unknown"


def test_the_item_number_wins_over_the_flavor():
    """Blinds have only ever shipped Dropship, but the column follows the item
    so that the day one ships DSD it is still reported as a blind."""
    assert product_for("DSD", {"vendor_items": ["138VB48D36WHTC"]}) == BLIND
    assert product_for("Dropship", {"vendor_items": ["72136-109-001"]}) == DRAPE


def _row(po, flavor, invoice):
    detail = {
        "metadata": {"id": "tx", "reference_id": invoice, "value": 100.0,
                     "source_document_reference_id": po,
                     "trading_partner_flavor": flavor},
        "file": {"generic_json_edi": {
            "heading": {"invoice_number": invoice, "invoice_date": "2026-09-01"},
            "detail": {"baseline_item_data_invoice_loop": [
                {"baseline_item_data_invoice": {"quantity_invoiced": "4", "unit_price": "25.00"}}]},
            "summary": {},
        }},
    }
    return detail


def test_extract_reads_the_product_off_the_po_index():
    po_index = {"PO-1": {"province": "ON", "vendor_items": ["138VB48D36WHTC"]},
                "PO-2": {"province": "QC", "vendor_items": ["72318-109-52-84-404"]}}
    assert extract(_row("PO-1", "Dropship", "INV1"), po_index)["product"] == BLIND
    assert extract(_row("PO-2", "Dropship", "INV2"), po_index)["product"] == DRAPE
    # DSD carries its own province and needs no PO entry to be placed
    assert extract(_row("PO-9", "Direct Store Delivery (DSD)", "INV3"), po_index)["product"] == DRAPE


def test_workbook_carries_product_next_to_type():
    po_index = {"PO-1": {"province": "ON", "vendor_items": ["138VB48D36WHTC"]}}
    rows = [extract(_row("PO-1", "Dropship", "INV1"), po_index)]
    ws = load_workbook(io.BytesIO(_bytes(rows)))["Invoices"]
    header = [c.value for c in ws[1]]
    assert header[:5] == ["Invoice", "Type", "Product", "Province", "Invoice Date"]
    assert [c.value for c in ws[2]][:3] == ["INV1", "Dropship", BLIND]


def test_total_row_stays_aligned_under_subtotal():
    """Product shifted every money column one to the right. If the total row's
    leading blanks are not shifted with them, the totals land under the wrong
    headings -- silently, since they are still numbers."""
    po_index = {"PO-1": {"province": "ON", "vendor_items": ["72136-109-001"]}}
    rows = [extract(_row("PO-1", "Dropship", "INV1"), po_index),
            extract(_row("PO-1", "Dropship", "INV2"), po_index)]
    ws = load_workbook(io.BytesIO(_bytes(rows)))["Invoices"]
    header = [c.value for c in ws[1]]
    total = [c.value for c in ws[ws.max_row]]
    assert total[0] == "Total"
    assert total[header.index("Subtotal")] == 200.0
    assert total[header.index("Total")] == 200.0


def _bytes(rows):
    buf = io.BytesIO()
    build_workbook(rows, "2026-09-01").save(buf)
    return buf.getvalue()


def _summary(rows):
    ws = load_workbook(io.BytesIO(_bytes(rows)))["Summary"]
    return [[c.value for c in r] for r in ws.iter_rows()]


def _block(grid, title):
    """The rows of one Summary block, from its header to its Total line."""
    start = next(i for i, r in enumerate(grid) if r[0] == title)
    end = next(i for i, r in enumerate(grid[start:], start) if r[0] == "Total")
    return grid[start + 1:end], grid[end]


def test_summary_splits_by_product_as_well_as_type():
    po_index = {"PO-B": {"province": "ON", "vendor_items": ["138VB48D36WHTC"]},
                "PO-D": {"province": "QC", "vendor_items": ["72318-109-52-84-404"]}}
    rows = [extract(_row("PO-B", "Dropship", "INV1"), po_index),
            extract(_row("PO-D", "Dropship", "INV2"), po_index),
            extract(_row("PO-D", "Dropship", "INV3"), po_index)]
    grid = _summary(rows)

    body, _ = _block(grid, "Product")
    assert [(r[0], r[1]) for r in body] == [(BLIND, 1), (DRAPE, 2)]


def test_the_two_summary_blocks_agree():
    """Type and Product are two cuts of the same invoices. If their totals ever
    diverge, one of the two columns is being read from something the other is
    not, and the sheet would be quietly contradicting itself."""
    po_index = {"PO-B": {"province": "ON", "vendor_items": ["138VB48D36WHTC"]},
                "PO-D": {"province": "QC", "vendor_items": ["72318-109-52-84-404"]}}
    rows = [extract(_row("PO-B", "Dropship", "INV1"), po_index),
            extract(_row("PO-D", "Direct Store Delivery (DSD)", "INV2"), po_index),
            extract(_row("PO-D", "Dropship", "INV3"), po_index)]
    grid = _summary(rows)

    by_type_body, by_type_total = _block(grid, "Type")
    by_product_body, by_product_total = _block(grid, "Product")
    assert by_type_total == by_product_total
    assert by_type_total[1] == 3                       # invoice count
    # and each block's rows add back up to its own total
    for body, total in ((by_type_body, by_type_total), (by_product_body, by_product_total)):
        for col in range(1, 7):
            assert round(sum(r[col] for r in body), 2) == total[col]


def test_summary_states_a_single_product_rather_than_omitting_the_block():
    """Most windows are drapery only. The block still appears, so the sheet
    says there were no blinds instead of leaving it unasked."""
    po_index = {"PO-D": {"province": "QC", "vendor_items": ["72318-109-52-84-404"]}}
    rows = [extract(_row("PO-D", "Dropship", "INV1"), po_index)]
    body, total = _block(_summary(rows), "Product")
    assert [(r[0], r[1]) for r in body] == [(DRAPE, 1)]
    assert total[1] == 1


def test_summary_money_is_formatted_in_both_blocks():
    po_index = {"PO-B": {"province": "ON", "vendor_items": ["138VB48D36WHTC"]}}
    rows = [extract(_row("PO-B", "Dropship", "INV1"), po_index)]
    ws = load_workbook(io.BytesIO(_bytes(rows)))["Summary"]
    money = [c.number_format for r in ws.iter_rows(min_row=3) for c in r
             if isinstance(c.value, (int, float)) and c.column > 2]
    assert money and set(money) == {"#,##0.00"}
    # the Invoices count is a count, not an amount
    counts = [c.number_format for r in ws.iter_rows(min_row=3) for c in r
              if isinstance(c.value, (int, float)) and c.column == 2]
    assert set(counts) == {"General"}
