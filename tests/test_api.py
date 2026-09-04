import io

import pytest
from openpyxl import load_workbook
from unittest.mock import patch, MagicMock
from fastapi.testclient import TestClient

from app.report import XLSX_MEDIA_TYPE

MOCK_INVOICES = [
    {
        "transaction_id": "tx-001",
        "invoice_number": "INV-001",
        "po_number": "PO-123",
        "trading_partner": "Home Depot",
        "invoice_date": "2026-07-01",
        "due_date": "2026-08-01",
        # Crstl returns "Accepted" or "Draft"; "Open" was invented here and
        # exists in no real payload. Reporting filters on Accepted, so a
        # fictional status made this fixture silently unreportable.
        "status": "Accepted",
        "subtotal": 4180.0,
        "tax_amount": 320.0,
        "total_amount": 4500.0,
        "currency": "USD",
        "created_at": "2026-07-01T00:00:00Z",
        "invoice_lines": [],
    }
]


def _raw_810(tx_id, subtotal, tax, *, invoice_date="2026-07-01",
             flavor="HD Canada Dropship", province="ON", po="PO-123"):
    """A minimal 810 detail payload in the shape app/report.py extracts from.

    The export re-reads each invoice from Crstl rather than trusting the cached
    figures, so a test that exercises /api/export has to stand up the raw
    payload too — a MagicMock detail would extract into a workbook of
    MagicMocks and assert nothing real.
    """
    return {
        "metadata": {
            "id": tx_id,
            "reference_id": f"INV-{tx_id}",
            "source_document_reference_id": po,
            "trading_partner_flavor": flavor,
            "state": {"value": "Accepted"},
            "value": round(subtotal + tax, 2),
            "created_at": f"{invoice_date}T00:00:00Z",
        },
        "file": {"generic_json_edi": {
            "heading": {
                "invoice_number": f"INV-{tx_id}",
                "invoice_date": invoice_date,
                "ship_to": {"state_province": province},
            },
            "detail": {"baseline_item_data_invoice_loop": [
                {"baseline_item_data_invoice": {
                    "quantity_invoiced": "1", "unit_price": f"{subtotal}"}}
            ]},
            # Dropship carries tax in TXI, not SAC. VA is HD's code for HST.
            "summary": {"tax_information": (
                [{"tax_type_code": "VA", "monetary_amount": f"{tax}"}] if tax else []
            )},
        }},
    }


# Amounts per transaction id, mirroring the cached fixtures below.
_RAW_AMOUNTS = {"tx-001": (4180.0, 320.0)}


def _detail_side_effect(tx_id):
    subtotal, tax = _RAW_AMOUNTS.get(tx_id, (100.0, 0.0))
    return _raw_810(tx_id, subtotal, tax)


def _invoice_rows(content):
    """Data rows of the workbook's Invoices sheet — header and Total excluded."""
    ws = load_workbook(io.BytesIO(content))["Invoices"]
    return [r for r in ws.iter_rows(min_row=2, values_only=True)
            if r and r[0] and r[0] != "Total"]


def _invoice_numbers(content):
    return [r[0] for r in _invoice_rows(content)]


def _invoice_header(content):
    """Header labels of the Invoices sheet.

    Assertions index by label rather than by position: the sheet has gained
    columns twice (Product, then Ship Date), and each time every positional
    assertion downstream of the insert broke without the figures themselves
    being wrong."""
    ws = load_workbook(io.BytesIO(content))["Invoices"]
    return [c.value for c in ws[1]]


@pytest.fixture
def client(monkeypatch, tmp_path):
    # Prevent .env MOCK_DATA=true from bypassing the patched CrstlClient
    monkeypatch.delenv("MOCK_DATA", raising=False)
    # Isolate tracking DB per test — must be set before TestClient starts the lifespan
    monkeypatch.setenv("TRACKING_DB", str(tmp_path / "tracking.db"))
    with patch("app.main.CrstlClient") as MockClient:
        mock_instance = MagicMock()
        mock_instance.fetch_invoices.return_value = MOCK_INVOICES
        # A real dict, not a MagicMock: _refresh_cache now keeps this map so the
        # export can resolve a Dropship province without re-crawling every 850.
        mock_instance.fetch_po_provinces.return_value = {}
        mock_instance._fetch_transaction_detail.side_effect = _detail_side_effect
        MockClient.return_value = mock_instance

        from app.main import app, _cache, _netsuite_state
        with TestClient(app) as c:
            # Reset cache after startup so each test controls its own state
            _cache["invoices"] = []
            _cache["po_provinces"] = {}
            _cache["last_synced"] = None
            _cache["status"] = "never"
            _netsuite_state["last_generated"] = None
            _netsuite_state["path"] = None
            _netsuite_state["count"] = 0
            _netsuite_state["skipped"] = 0
            _netsuite_state["error"] = None
            _netsuite_state["generating"] = False
            yield c


def test_get_invoices_returns_list(client):
    resp = client.get("/api/invoices")
    assert resp.status_code == 200
    data = resp.json()
    assert "invoices" in data
    assert "last_synced" in data
    assert "status" in data


def test_sync_triggers_refresh(client):
    resp = client.post("/api/sync")
    assert resp.status_code == 200
    data = resp.json()
    assert data["ok"] is True
    assert data["last_synced"] is not None
    assert data["status"] == "ok"


def test_export_all_returns_xlsx(client):
    # Pre-populate cache via sync
    client.post("/api/sync")
    resp = client.post("/api/export", json={})
    assert resp.status_code == 200
    assert resp.headers["content-type"] == XLSX_MEDIA_TYPE
    assert "attachment" in resp.headers["content-disposition"]
    assert resp.headers["content-disposition"].endswith('.xlsx"')
    assert _invoice_numbers(resp.content) == ["INV-tx-001"]


def test_export_subset_by_ids(client):
    client.post("/api/sync")
    resp = client.post("/api/export", json={"ids": ["tx-001"]})
    assert resp.status_code == 200
    assert _invoice_numbers(resp.content) == ["INV-tx-001"]


def test_export_reports_the_810_figures_not_the_cached_ones(client):
    """The cache infers Dropship tax from a province rate table; the workbook
    reads the TXI segment HD actually sent. This asserts the workbook is built
    from the payload, so the two can never silently diverge behind one
    filename."""
    client.post("/api/sync")
    resp = client.post("/api/export", json={"ids": ["tx-001"]})
    row = _invoice_rows(resp.content)[0]
    at = _invoice_header(resp.content).index
    assert row[at("Type")] == "Dropship"
    # fetch_po_provinces is stubbed empty here, so there is no 850 to read the
    # ordered items from. A Dropship invoice with no PO on file reports
    # Unknown rather than defaulting to the product line we sell more of.
    assert row[at("Product")] == "Unknown"
    assert row[at("Province")] == "ON"
    assert row[at("Subtotal")] == 4180.0    # subtotal from the line loop
    assert row[at("Tax")] == 320.0          # tax from TXI, not from a rate table
    assert row[at("Total")] == 4500.0       # metadata.value, HD's stated total


def test_export_502s_when_crstl_returns_nothing(client):
    """An empty workbook reads as "a quiet day" to whoever opens it. If every
    detail fetch failed, that must surface as an error instead."""
    from app.main import _cache
    client.post("/api/sync")
    with patch("app.main.rows_for_transactions", return_value=[]):
        resp = client.post("/api/export", json={})
    assert resp.status_code == 502
    assert "no detail" in resp.json()["message"]


def test_export_empty_cache_returns_503(client):
    # Don't sync — cache is empty
    resp = client.post("/api/export", json={})
    assert resp.status_code == 503


def test_netsuite_returns_501(client):
    resp = client.post("/api/netsuite")
    assert resp.status_code == 501
    assert "not yet configured" in resp.json()["message"]


def test_export_sets_exported_at(client):
    client.post("/api/sync")
    client.post("/api/export", json={})

    resp = client.get("/api/invoices")
    invoices = resp.json()["invoices"]
    assert len(invoices) > 0
    assert invoices[0]["exported_at"] is not None
    assert invoices[0]["netsuite_at"] is None


def test_netsuite_export_latest_initially_unavailable(client):
    resp = client.get("/api/netsuite-export/latest")
    assert resp.status_code == 200
    data = resp.json()
    assert data["available"] is False
    assert data["last_generated"] is None
    assert data["count"] == 0


def test_netsuite_export_download_no_file_returns_404(client):
    resp = client.get("/api/netsuite-export/download")
    assert resp.status_code == 404


def test_netsuite_generate_and_download(client, monkeypatch, tmp_path):
    monkeypatch.setenv("MOCK_DATA", "true")
    client.post("/api/sync")
    resp = client.post("/api/netsuite-export/generate")
    assert resp.status_code == 200
    assert resp.json()["available"] is True

    download = client.get("/api/netsuite-export/download")
    assert download.status_code == 200
    assert "text/csv" in download.headers["content-type"]
    assert "netsuite_export" in download.headers["content-disposition"]


def _scheduled_jobs(monkeypatch, tmp_path):
    """Start the app lifespan with a stubbed scheduler and return the cron
    kwargs each job registered with, keyed by job id."""
    monkeypatch.delenv("MOCK_DATA", raising=False)
    monkeypatch.setenv("TRACKING_DB", str(tmp_path / "tracking.db"))
    monkeypatch.setenv("SCHEDULER_ENABLED", "true")
    jobs: dict[str, dict] = {}

    with patch("app.main.CrstlClient") as MockClient, \
         patch("app.main.AsyncIOScheduler") as MockScheduler:
        mock_instance = MagicMock()
        mock_instance.fetch_invoices.return_value = MOCK_INVOICES
        MockClient.return_value = mock_instance

        def add_job(func, trigger, **kwargs):
            jobs[kwargs["id"]] = {"trigger": trigger, **kwargs}

        MockScheduler.return_value.add_job.side_effect = add_job

        from app.main import app
        with TestClient(app):
            pass

    return jobs


def test_digest_scheduled_weekdays_only(monkeypatch, tmp_path):
    """The digest must not fire Sat/Sun — Friday's late invoices ride along
    with Monday's send because tracking.db still lists them as unemailed."""
    jobs = _scheduled_jobs(monkeypatch, tmp_path)
    digest = jobs["daily_digest"]
    assert digest["day_of_week"] == "mon-fri"
    assert (digest["hour"], digest["minute"]) == (7, 15)
    assert digest["timezone"] == "America/Toronto"


def test_refresh_and_export_still_run_every_day(monkeypatch, tmp_path):
    """Only the email is weekday-gated; the cache refresh and NetSuite export
    keep running on weekends so Monday's digest has current data."""
    jobs = _scheduled_jobs(monkeypatch, tmp_path)
    for job_id in ("daily_refresh", "netsuite_export"):
        assert "day_of_week" not in jobs[job_id]


# ── Accepted-only reporting ────────────────────────────────────────────────
# The warehouse resubmits invoices when it catches a mistake, and each attempt
# lands as its own Crstl record. Only the acknowledged one may reach accounting.

def _mixed_status_invoices():
    def inv(tx, num, status, total):
        return {
            "transaction_id": tx, "invoice_number": num, "po_number": "PO-1",
            "trading_partner": "Home Depot Canada", "province": "ON",
            "invoice_date": "2026-08-24", "due_date": "2026-09-23",
            "status": status, "subtotal": total, "tax_amount": 0.0,
            "total_amount": total, "currency": "CAD",
            "created_at": "2026-08-24T00:00:00Z", "invoice_lines": [],
            "allowances_charges": [],
        }
    # one accepted invoice plus two superseded drafts of the same PO
    return [inv("tx-acc", "INV-1", "Accepted", 100.0),
            inv("tx-dr1", "INV-1", "Draft", 100.0),
            inv("tx-dr2", "INV-1", "Draft", 100.0)]


def test_reportable_keeps_accepted_and_drops_drafts():
    from app.main import _reportable
    kept = _reportable(_mixed_status_invoices())
    assert [i["transaction_id"] for i in kept] == ["tx-acc"]


def test_reportable_withholds_unknown_status_and_warns(capsys):
    """An unrecognised state must not reach accounting silently. If Crstl ever
    adds a status that means "good", this warning is how we find out rather
    than invoices quietly vanishing from the digest."""
    from app.main import _reportable
    rows = _mixed_status_invoices() + [{"transaction_id": "tx-new", "status": "Submitted"}]
    kept = _reportable(rows)
    assert [i["transaction_id"] for i in kept] == ["tx-acc"]
    out = capsys.readouterr().out
    assert "Submitted" in out and "REPORTABLE_STATUSES" in out
    # a plain Draft is expected and must not generate noise
    assert "Draft" not in out


def test_bulk_export_excludes_drafts(client):
    """A bulk export is a report, so it carries Accepted only."""
    from app.main import _cache
    client.post("/api/sync")
    with patch.dict(_cache, {"invoices": _mixed_status_invoices()}):
        content = client.post("/api/export", json={}).content
    numbers = _invoice_numbers(content)
    assert numbers == ["INV-tx-acc"], f"expected only the Accepted invoice, got {numbers}"


def test_explicit_id_selection_is_honoured_even_for_a_draft(client):
    """Picking specific invoices on the dashboard is a deliberate act, so it is
    not second-guessed — only the automatic bulk report filters."""
    from app.main import _cache
    client.post("/api/sync")
    with patch.dict(_cache, {"invoices": _mixed_status_invoices()}):
        content = client.post("/api/export", json={"ids": ["tx-dr1"]}).content
    assert _invoice_numbers(content) == ["INV-tx-dr1"]


def test_draft_is_not_marked_emailed_so_it_can_be_reported_once_accepted(client):
    """Filtering defers, it must never lose. A Draft withheld today has to
    still be unemailed tomorrow, so the digest picks it up the moment HD
    acknowledges it."""
    from app.main import _cache, _send_daily_digest
    from app.tracking import get_unemailed_ids
    client.post("/api/sync")
    with patch.dict(_cache, {"invoices": _mixed_status_invoices()}), \
         patch("app.main.send_mail") as mail:
        result = _send_daily_digest()
    assert result["count"] == 1, "only the Accepted invoice should be emailed"
    assert mail.called
    # the drafts remain unemailed and are therefore still eligible later
    assert set(get_unemailed_ids(["tx-acc", "tx-dr1", "tx-dr2"])) == {"tx-dr1", "tx-dr2"}


def test_digest_attaches_the_workbook(client, monkeypatch):
    """The digest carries the same workbook the Export button produces. These
    are the two surfaces accounting actually receives, so they must not drift
    into different formats — or different figures — from one another."""
    from app.main import _send_daily_digest
    monkeypatch.setenv("MAIL_RECIPIENTS", "accounting@example.com")
    client.post("/api/sync")
    with patch("app.main.send_mail") as mail:
        result = _send_daily_digest()
    assert result["count"] == 1
    attachments = mail.call_args.kwargs["attachments"]
    assert len(attachments) == 1
    name, content, mime = attachments[0]
    assert name.endswith(".xlsx")
    assert mime == XLSX_MEDIA_TYPE
    assert _invoice_numbers(content) == ["INV-tx-001"]


def test_digest_aborts_rather_than_emailing_without_the_workbook(client, monkeypatch):
    """If Crstl can't be reached the digest must fail, not send. Nothing is
    marked emailed, so tomorrow's digest still carries these invoices instead
    of accounting receiving a mail that quietly omits them."""
    from app.main import _send_daily_digest, ReportUnavailable
    from app.tracking import get_unemailed_ids
    monkeypatch.setenv("MAIL_RECIPIENTS", "accounting@example.com")
    client.post("/api/sync")
    with patch("app.main.rows_for_transactions", return_value=[]), \
         patch("app.main.send_mail") as mail:
        with pytest.raises(ReportUnavailable):
            _send_daily_digest()
    assert not mail.called
    assert get_unemailed_ids(["tx-001"]) == ["tx-001"]
