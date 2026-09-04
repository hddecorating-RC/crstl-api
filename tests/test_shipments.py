"""The Ship Date and Pickup Date columns, read from the 856.

The shapes asserted here are real: Dropship and Wholesale ASNs spell the field
`shipped_date`, DSD spells it `shipment_date`, and Wholesale's heading-level
`ship_notice_date` is a 2008 stub on every ASN on record. A test that read the
heading date would pass on DSD and quietly date every Wholesale line to 2008.
"""
import io

from openpyxl import load_workbook

from app.report import build_workbook, extract
from app.shipments import (asn_date_index, asn_date_label, asn_dates_in,
                           merge_asn_dates)


def _asn(po, dates, field="shipped_date", state="Accepted", notice="2026-08-18"):
    return {
        "metadata": {"id": f"tx-{po}", "reference_id": f"ASN{po}",
                     "source_document_reference_id": po,
                     "state": {"value": state}},
        "file": {"generic_json_edi": {
            "heading": {"ship_notice_date": notice},
            "detail": {"shipments": [{field: d} for d in dates]},
        }},
    }


def test_dropship_and_dsd_spell_the_field_differently():
    assert asn_dates_in(_asn("1", ["2026-08-18"], "shipped_date")) == ["2026-08-18"]
    assert asn_dates_in(_asn("1", ["2026-06-29"], "shipment_date")) == ["2026-06-29"]


def test_the_heading_notice_date_is_never_read():
    """Every Wholesale ASN carries ship_notice_date 2008-11-10. Falling back to
    it would date those lines to 2008 instead of leaving them blank."""
    asn = _asn("90000000", [], notice="2008-11-10")
    assert asn_dates_in(asn) == []


def test_dates_are_distinct_and_sorted():
    asn = _asn("1", ["2026-03-26", "2026-03-24", "2026-03-24"])
    assert asn_dates_in(asn) == ["2026-03-24", "2026-03-26"]


def test_a_single_date_is_shown_plainly():
    assert asn_date_label(["2026-08-18"]) == "2026-08-18"


def test_a_split_shipment_is_shown_as_its_range():
    """3 POs on record carry more than one Accepted ASN. Showing only the first
    would date the invoice earlier than half of what it covers."""
    assert asn_date_label(["2026-03-24", "2026-03-26"]) == "2026-03-24 to 2026-03-26"


def test_no_asn_is_blank_not_a_guess():
    assert asn_date_label([]) == ""


def test_drafts_are_superseded_resubmissions_and_do_not_count():
    """9 POs carry an Accepted ASN alongside a Draft one. Both spell the same
    date today, but a Draft is a resubmission that was never sent."""
    index = asn_date_index([_asn("40825462", ["2026-03-24"], state="Accepted"),
                            _asn("40825462", ["2026-03-30"], state="Draft")])
    assert index == {"40825462": "2026-03-24"}


def test_rejected_asns_do_not_count_either():
    index = asn_date_index([_asn("P1", ["2026-03-24"], state="Rejected")])
    assert index == {}


def test_two_accepted_asns_on_one_po_merge_into_a_range():
    index = asn_date_index([_asn("P1", ["2026-03-24"]), _asn("P1", ["2026-03-26"])])
    assert index == {"P1": "2026-03-24 to 2026-03-26"}


def _invoice(po, invoice="INV1", flavor="Dropship"):
    return {
        "metadata": {"id": "tx", "reference_id": invoice, "value": 100.0,
                     "source_document_reference_id": po,
                     "trading_partner_flavor": flavor},
        "file": {"generic_json_edi": {
            "heading": {"invoice_number": invoice, "invoice_date": "2026-09-01"},
            "detail": {"baseline_item_data_invoice_loop": [
                {"baseline_item_data_invoice": {"quantity_invoiced": "4", "unit_price": "25.00"}}]},
            "summary": {},
        }},
    }


def test_dropship_asn_date_is_a_real_ship_date():
    po_index = {"PO-1": {"province": "ON", "vendor_items": ["72136-109-001"],
                         "asn_date": "2026-08-18"}}
    row = extract(_invoice("PO-1"), po_index)
    assert row["ship_date"] == "2026-08-18"
    assert row["pickup_date"] == ""


def test_dsd_asn_date_is_a_pickup_date_and_never_a_ship_date():
    """HD requires a pickup date to raise a DSD ASN; the goods leave 3-4 days
    later. Putting it under Ship Date would be wrong by 3-4 days on every DSD
    line and would read as data rather than as the gap it is. The actual date
    is in Finale, which this service does not read."""
    po_index = {"PO-1": {"province": "ON", "asn_date": "2026-09-08"}}
    row = extract(_invoice("PO-1", flavor="Direct Store Delivery (DSD)"), po_index)
    assert row["pickup_date"] == "2026-09-08"
    assert row["ship_date"] == ""


def test_an_invoice_with_no_asn_leaves_both_cells_blank():
    po_index = {"PO-1": {"province": "ON", "vendor_items": ["72136-109-001"]}}
    row = extract(_invoice("PO-1"), po_index)
    assert row["ship_date"] == "" and row["pickup_date"] == ""


def _bytes(rows):
    buf = io.BytesIO()
    build_workbook(rows, "2026-09-01").save(buf)
    return buf.getvalue()


def test_workbook_carries_both_date_columns_after_invoice_date():
    po_index = {"PO-1": {"province": "ON", "vendor_items": ["72136-109-001"],
                         "asn_date": "2026-08-18"}}
    rows = [extract(_invoice("PO-1"), po_index)]
    ws = load_workbook(io.BytesIO(_bytes(rows)))["Invoices"]
    assert [c.value for c in ws[1]][:7] == ["Invoice", "Type", "Product", "Province",
                                            "Invoice Date", "Ship Date", "Pickup Date"]
    row = [c.value for c in ws[2]]
    assert row[:6] == ["INV1", "Dropship", "Drape Panel", "ON",
                       "2026-09-01", "2026-08-18"]
    assert row[6] in (None, "")          # openpyxl reads an empty cell as None


def test_a_dsd_row_fills_pickup_and_leaves_ship_blank_in_the_sheet():
    po_index = {"PO-1": {"province": "ON", "asn_date": "2026-09-08"}}
    rows = [extract(_invoice("PO-1", flavor="Direct Store Delivery (DSD)"), po_index)]
    ws = load_workbook(io.BytesIO(_bytes(rows)))["Invoices"]
    at = [c.value for c in ws[1]].index
    row = [c.value for c in ws[2]]
    assert row[at("Pickup Date")] == "2026-09-08"
    assert row[at("Ship Date")] in (None, "")


def test_total_row_stays_aligned_after_the_dates_shift_the_money():
    """Ship Date shifts every money column one further right. If the total
    row's leading blanks are not shifted with them the totals land under the
    wrong headings, silently, since they are still numbers."""
    po_index = {"PO-1": {"province": "ON", "vendor_items": ["72136-109-001"],
                         "asn_date": "2026-08-18"}}
    rows = [extract(_invoice("PO-1", "INV1"), po_index),
            extract(_invoice("PO-1", "INV2"), po_index)]
    ws = load_workbook(io.BytesIO(_bytes(rows)))["Invoices"]
    header = [c.value for c in ws[1]]
    total = [c.value for c in ws[ws.max_row]]
    assert total[0] == "Total"
    assert total[header.index("Subtotal")] == 200.0
    assert total[header.index("Total")] == 200.0
    assert total[header.index("Ship Date")] in (None, "")
    assert total[header.index("Pickup Date")] in (None, "")


def test_merge_keeps_what_the_850_already_said():
    po_index = {"P1": {"province": "ON", "vendor_items": ["72136-109-001"]}}
    merge_asn_dates(po_index, {"P1": "2026-08-18"})
    assert po_index["P1"] == {"province": "ON", "vendor_items": ["72136-109-001"],
                              "asn_date": "2026-08-18"}


def test_merge_records_a_shipment_whose_850_never_arrived():
    """The Ship Date is real even when the PO is not on file. Dropping it would
    hide a shipment because a different document is missing."""
    po_index = {}
    merge_asn_dates(po_index, {"P9": "2026-08-18"})
    assert po_index == {"P9": {"asn_date": "2026-08-18"}}
