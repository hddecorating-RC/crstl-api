"""The daily invoice workbook: one row per invoice, plus a Summary sheet.

Shared by the app -- the dashboard's Export button and the digest email
attachment -- and by tools/daily_invoice_report_xlsx.py, which is a CLI over
these same functions. The columns accounting reads are defined here once.

Rows are built from the raw 810 payload rather than from the cached invoice
dicts app/crstl.py produces, because the two disagree on tax in ways that
reach the total:

  * The indicator decides an entry's sign, not the code. app/sac_codes.py
    classify() keys on the code first, so a D360 sent with indicator "A"
    lands in tax instead of allowances.
  * Dropship and Wholesale carry tax in TXI (summary.tax_information), which
    the cached path does not read at all -- app/main.py infers it from a
    province rate table instead.

Tax reaches us by two routes depending on the flavor, and both are read:

    Dropship / Wholesale   tax_information (TXI)   CG=GST  ST=QST  VA=HST
    DSD                    SAC codes               D360=GST  H680=QST
                                                   H770=HST  H850=ECO

That SAC reading follows the HD revision confirmed with Crstl in Aug 2026 and
HD's own rejection of INV40852200, not the v6.10 PDF in this repo, which
predates H770 and routes HST to H680.

Dropship invoices carry no ship-to, so the province is recovered from the 850
PO; DSD carries its own.

Ship Date and Pickup Date are two columns because the 856 sends one date that
means two different things: the real ship date on Dropship and Wholesale, a
scheduled pickup on DSD whose goods leave 3-4 days later. A DSD Ship Date is
left blank rather than filled with the pickup date -- the actual one is in
Finale, which this service does not read. app/shipments.py has the evidence.

The Product column -- drapery or blinds -- comes from the 850 as well. An 810
carries no vendor item number on any flavor, and the 850 is already fetched
for the province, so the same po_index answers both. app/products.py holds the
rule; DSD is settled here, because DSD 850s carry no item number either.
"""
import io
from concurrent.futures import ThreadPoolExecutor, as_completed

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.crstl import CrstlClient
from app.products import DRAPE, UNKNOWN, label_for


num = CrstlClient._parse_float             # strips thousands separators

# Tax vocabularies. Crstl's own UI spells CG out as "Federal Value Added Tax
# GST On Goods"; these readings come from Crstl and from HD, not from matching
# amounts against province rates.
TXI_KIND = {"CG": "GST", "ST": "QST", "VA": "HST"}
SAC_KIND = {"D360": "GST", "H680": "QST", "H770": "HST",
            "H850": "ECO", "F240": "ECO", "G090": "ECO", "G100": "ECO"}

MONEY = "#,##0.00"
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
BAD_FILL = PatternFill("solid", fgColor="FFC7CE")

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


# Matches app/main.py's rate-check convention: a 2-cent floor, widening to 0.2%
# on larger invoices. HD rounds each line to the cent, so summing lines can
# drift a cent or two from the stated total without anything being wrong.
def tolerance_for(amount):
    return max(0.02, round(0.002 * abs(amount), 2))


def flavor_of(meta):
    """Crstl sends three flavors. Wholesale is rare but real, and collapsing it
    into Dropship would mislabel the column accounting reads to tell the
    channels apart."""
    raw = str(meta.get("trading_partner_flavor") or "")
    if "DSD" in raw or "Direct Store" in raw:
        return "DSD"
    if "Wholesale" in raw:
        return "Wholesale"
    if "Dropship" in raw:
        return "Dropship"
    return raw or "Unknown"


def product_for(flavor, po_entry):
    """The Product column for one invoice: what was sold, not how it shipped.

    Blinds have only ever reached us by Dropship, but the Dropship reading is
    keyed on the item number rather than on the flavor, so the day a blind
    ships another way the column follows the item. DSD is the one flavor read
    from the flavor itself -- its 850s carry no item number to read instead --
    and every DSD item HD has ordered is drapery.
    """
    label = label_for((po_entry or {}).get("vendor_items") or ())
    if label:
        return label
    if flavor == "DSD":
        return DRAPE
    return UNKNOWN


def dates_for(flavor, asn_date):
    """Split the one date the 856 carries into the two columns it means.

    Dropship and Wholesale send the real ship date. DSD sends a scheduled
    pickup date -- HD requires one to raise the ASN -- and the goods ship 3-4
    days later; that actual date is in Finale, not in any EDI document we
    receive. app/shipments.py has the evidence.

    So a DSD row's Ship Date is left blank on purpose. Putting the pickup date
    there would read as a ship date to accounting and be wrong by 3-4 days on
    every DSD line, which is worse than saying nothing. The blank is the slot
    Finale fills.
    """
    if flavor == "DSD":
        return {"ship_date": "", "pickup_date": asn_date}
    return {"ship_date": asn_date, "pickup_date": ""}


def extract(detail, po_index):
    """Flatten one 810 payload into the row accounting needs."""
    meta = detail.get("metadata", {}) or {}
    edi = (detail.get("file", {}) or {}).get("generic_json_edi", {}) or {}
    heading = edi.get("heading", {}) or {}
    summary = edi.get("summary", {}) or {}

    po = str(meta.get("source_document_reference_id") or heading.get("purchase_order_number") or "")
    po_entry = po_index.get(po) or {}
    ship_to = heading.get("ship_to") or {}
    province = (str(ship_to.get("state_province") or "").upper()
                or str(po_entry.get("province") or "").upper())
    flavor = flavor_of(meta)

    subtotal = round(sum(
        num((e.get("baseline_item_data_invoice", {}) or {}).get("quantity_invoiced"))
        * num((e.get("baseline_item_data_invoice", {}) or {}).get("unit_price"))
        for e in (edi.get("detail", {}) or {}).get("baseline_item_data_invoice_loop", []) or []
    ), 2)

    deductions, charges, taxes, unknown = {}, {}, {}, {}
    for entry in summary.get("service_promotion_allowance_or_charge_information_loop", []) or []:
        sac = entry.get("service_promotion_allowance_or_charge_information", {}) or {}
        code = sac.get("service_promotion_allowance_or_charge_code") or "(no code)"
        amt = num(sac.get("amount"))
        indicator = sac.get("allowance_or_charge_indicator")
        # The indicator decides, not the code. HD has sent D360 -- a tax code --
        # with indicator "A" on 3 invoices, and reading it as tax puts those
        # 70.52 on the wrong side: INV40831749 reconciles at 5,317.43 only if
        # the allowance is honoured as an allowance. Report what was sent.
        if indicator == "A":
            deductions[code] = round(deductions.get(code, 0) + amt, 2)
        elif indicator == "C":
            kind = SAC_KIND.get(code)
            if kind:
                taxes[kind] = round(taxes.get(kind, 0) + amt, 2)
            else:
                charges[code] = round(charges.get(code, 0) + amt, 2)
        else:
            # Neither allowance nor charge. Its sign is unknowable, so it is
            # shown but kept out of the reconciliation rather than guessed at.
            unknown[code] = round(unknown.get(code, 0) + amt, 2)

    for entry in summary.get("tax_information", []) or []:
        kind = TXI_KIND.get(entry.get("tax_type_code")) or entry.get("tax_type_code") or "(no code)"
        taxes[kind] = round(taxes.get(kind, 0) + num(entry.get("monetary_amount")), 2)

    # metadata.value is Crstl's canonical total (app/crstl.py says so); fall
    # back to the EDI summary only when it is genuinely absent, not when it is
    # a legitimate 0.00.
    stated = num(meta.get("value") if meta.get("value") is not None
                 else (summary.get("total_monetary_value_summary", {}) or {}).get("total_amount"))

    # An 810 with no parsable line loop reports subtotal 0; the app treats the
    # stated total as the subtotal in that case rather than under-reporting.
    if subtotal == 0.0:
        subtotal = stated

    computed = round(subtotal - sum(deductions.values()) + sum(charges.values())
                     + sum(taxes.values()), 2)

    return {
        "transaction_id": str(meta.get("id") or detail.get("id") or ""),
        "invoice": str(meta.get("reference_id") or heading.get("invoice_number") or ""),
        "date": heading.get("invoice_date") or "",
        "po": po,
        "flavor": flavor,
        "product": product_for(flavor, po_entry),
        "province": province,
        **dates_for(flavor, str(po_entry.get("asn_date") or "")),
        "subtotal": subtotal,
        "deductions": deductions,
        "charges": charges,
        "taxes": taxes,
        "unknown": unknown,
        "stated": stated,
        "computed": computed,
        "variance": round(stated - computed, 2),
    }


def reconciles(row):
    return abs(row["variance"]) <= tolerance_for(row["stated"])


def build_workbook(rows, window):
    """Summary columns, not a column per code. Returns an openpyxl Workbook.

    Type is how the invoice shipped, Product is what was on it -- an invoice
    can be Dropship and carry either line, so accounting needs both.

    Charges and Unclassified are included only when some invoice in the window
    actually carries one -- freight has never been transmitted, so on a normal
    day the sheet is the plain seven columns, and the moment a freight or fee
    amount arrives it appears as its own column instead of being folded into a
    total that then does not add up.
    """
    has_charges = any(r["charges"] for r in rows)
    has_unknown = any(r["unknown"] for r in rows)

    def tot(rs, key):
        return round(sum(sum(r[key].values()) for r in rs), 2)

    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"

    headers = (["Invoice", "Type", "Product", "Province", "Invoice Date",
                "Ship Date", "Pickup Date", "Subtotal", "Discounts"]
               + (["Charges"] if has_charges else [])
               + ["Tax"]
               + (["Unclassified"] if has_unknown else [])
               + ["Total"])
    ws.append(headers)
    for cell in ws[1]:
        cell.fill, cell.font = HDR_FILL, HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r in rows:
        row = [r["invoice"], r["flavor"], r["product"], r["province"], r["date"],
               r["ship_date"], r["pickup_date"], r["subtotal"],
               -round(sum(r["deductions"].values()), 2) or 0]
        if has_charges:
            row.append(round(sum(r["charges"].values()), 2) or 0)
        row.append(round(sum(r["taxes"].values()), 2) or 0)
        if has_unknown:
            row.append(round(sum(r["unknown"].values()), 2) or 0)
        row.append(r["stated"])
        ws.append(row)

    last = ws.max_row + 1
    total_row = ["Total", "", "", "", "", "", "",
                 round(sum(r["subtotal"] for r in rows), 2),
                 -tot(rows, "deductions")]
    if has_charges:
        total_row.append(tot(rows, "charges"))
    total_row.append(tot(rows, "taxes"))
    if has_unknown:
        total_row.append(tot(rows, "unknown"))
    total_row.append(round(sum(r["stated"] for r in rows), 2))
    ws.append(total_row)
    for cell in ws[last]:
        cell.font = Font(bold=True)
        cell.border = Border(top=Side(style="thin"))

    # Subtotal is column 8 now that Ship Date and Pickup Date sit between the
    # dates and the money. Formatting earlier would money-format the dates.
    for row in ws.iter_rows(min_row=2, min_col=8):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = MONEY
    widths = [18, 11, 13, 10, 14, 16, 16, 13] + [12] * (len(headers) - 9) + [13]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last - 1}"

    s = wb.create_sheet("Summary")
    s.append(["Invoiced", window]); s["A1"].font = Font(bold=True, size=13)
    s.append(["Source", "Crstl 810 transactions, state = Accepted"])

    def block(title, key):
        """One breakdown of the money, split by `key`.

        Type is how an invoice shipped, Product is what was on it. They are
        two cuts of the same invoices, not two subsets, so each block carries
        its own Total row and the two must agree -- a reader who adds up one
        block can check it against the other without leaving the sheet.

        A block is written even when everything falls in one bucket. Drapery
        outsells blinds by an order of magnitude, so most windows have a
        single Product row, and a "Drape Panel 52" line that equals the total
        is the sheet stating there were no blinds rather than staying silent
        about them.
        """
        s.append([])
        s.append([title, "Invoices", "Subtotal", "Discounts", "Charges", "Tax", "Total"])
        for cell in s[s.max_row]:
            cell.fill, cell.font = HDR_FILL, HDR_FONT
        for value in sorted({r[key] for r in rows}):
            sub = [r for r in rows if r[key] == value]
            s.append([value, len(sub), round(sum(r["subtotal"] for r in sub), 2),
                      -tot(sub, "deductions"), tot(sub, "charges"), tot(sub, "taxes"),
                      round(sum(r["stated"] for r in sub), 2)])
        s.append(["Total", len(rows), round(sum(r["subtotal"] for r in rows), 2),
                  -tot(rows, "deductions"), tot(rows, "charges"), tot(rows, "taxes"),
                  round(sum(r["stated"] for r in rows), 2)])
        for cell in s[s.max_row]:
            cell.font = Font(bold=True)

    block("Type", "flavor")
    block("Product", "product")

    unbalanced = [r for r in rows if not reconciles(r)]
    s.append([])
    s.append(["Do not match HD's total", len(unbalanced)])
    s[s.max_row][0].font = Font(bold=True)
    if unbalanced:
        s[s.max_row][1].fill = BAD_FILL
        s.append(["Invoice", "HD total", "Adds up to", "Difference"])
        for cell in s[s.max_row]:
            cell.font = Font(bold=True)
        for r in unbalanced:
            s.append([r["invoice"], r["stated"], r["computed"], r["variance"]])
    for row in s.iter_rows(min_row=3):
        for cell in row:
            if isinstance(cell.value, (int, float)) and cell.column > 2:
                cell.number_format = MONEY
    for i, w in enumerate((24, 14, 14, 14, 12, 12, 14), start=1):
        s.column_dimensions[get_column_letter(i)].width = w

    return wb


def workbook_bytes(rows, window):
    """The workbook as bytes, for an HTTP response or a mail attachment."""
    buf = io.BytesIO()
    build_workbook(rows, window).save(buf)
    return buf.getvalue()


def save_workbook(rows, out_path, window):
    build_workbook(rows, window).save(out_path)
    return out_path


def window_label(rows):
    """The Summary sheet's "Invoiced" line. An export is not date-windowed the
    way the CLI is, so the label is read back off the rows themselves."""
    dates = sorted(r["date"] for r in rows if r["date"])
    if not dates:
        return "date not stated"
    return dates[0] if dates[0] == dates[-1] else f"{dates[0]} to {dates[-1]}"


def fetch_details(client, transactions):
    """Fetch details in parallel, skipping records that fail rather than losing
    the whole report to one bad one -- the same trade-off fetch_invoices makes.

    Accepts either transaction records from a listing or bare id strings.
    """
    out = []
    with ThreadPoolExecutor(max_workers=CrstlClient.MAX_WORKERS) as pool:
        futures = {}
        for t in transactions:
            if isinstance(t, str):
                tid, label = t, t
            else:
                tid = t.get("id") or (t.get("metadata") or {}).get("id")
                label = t.get("reference_id")
            if not tid:
                print(f"WARNING: skipping transaction with no id: {label}")
                continue
            futures[pool.submit(client._fetch_transaction_detail, tid)] = label
        for fut in as_completed(futures):
            try:
                out.append(fut.result())
            except Exception as exc:
                print(f"WARNING: could not fetch {futures[fut]}: {exc}")
    return out


def rows_for_transactions(client, transaction_ids, po_index):
    """Rows for a known set of transaction ids, in the workbook's sort order.

    Used by the app, which already knows which invoices it is exporting, so it
    skips the CLI's listing crawl. Rows are not filtered by invoice_date here:
    an export carries what the caller asked for, and dropping an undated
    invoice would silently shrink a deliberate selection.
    """
    rows = [extract(d, po_index) for d in fetch_details(client, list(transaction_ids))]
    rows.sort(key=lambda r: (r["flavor"], r["invoice"]))
    return rows
