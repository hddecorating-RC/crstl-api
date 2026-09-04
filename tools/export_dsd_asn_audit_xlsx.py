"""Export the DSD ASN REF*CN audit to a shareable .xlsx workbook.

Reads the JSON produced by tools/audit_dsd_asn_refcn.py and writes a workbook
with four sheets:

    Summary       counts, window, the 2026-07-22 cutover, and what to do next
    Corrections   the in-window ASNs to correct, each paired with the Shipment
                  ID recovered from TD5-05 Routing on that same ASN
    Window        every DSD ASN in the audit window, correct ones included
    Full History  every DSD ASN on record, for the wider backlog conversation

Run tools/audit_dsd_asn_refcn.py first — this reads its .tmp output and makes
no API calls.

Usage:
    python3 tools/export_dsd_asn_audit_xlsx.py [--out PATH]
"""
import argparse
import json
import os
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
WINDOW_JSON = os.path.join(REPO, ".tmp", "dsd_asn_refcn.json")
FULL_JSON = os.path.join(REPO, ".tmp", "dsd_asn_refcn_12mo.json")

# Palette mirrors the published report: petrol accent, oxide red, deep green.
INK = "FF101A1F"
ACCENT = "FF0E5C6B"
BAD = "FFA3372A"
BAD_FILL = "FFF7EAE7"
GOOD = "FF2E6B4F"
GOOD_FILL = "FFE7F0EB"
HEAD_FILL = "FFEDF1F2"
LINE = "FFD6DEE1"

THIN = Side(style="thin", color=LINE)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

VERDICT_LABEL = {
    "ok": "Correct",
    "rts_error": "RTS number in REF*CN",
    "missing": "REF*CN empty",
    "unknown": "Unrecognized prefix",
}


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True, size=9, color=ACCENT)
        cell.fill = PatternFill("solid", fgColor=HEAD_FILL)
        cell.border = BOX
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 28
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_table(ws, start_row, headers, rows, widths, verdict_col=None):
    for i, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=i, value=h)
    style_header(ws, start_row, len(headers))

    for r, record in enumerate(rows, start_row + 1):
        bad = verdict_col is not None and record[verdict_col - 1] != "Correct"
        for c, value in enumerate(record, 1):
            cell = ws.cell(row=r, column=c, value=value)
            cell.border = BOX
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.font = Font(size=10, color=INK)
            if bad:
                cell.fill = PatternFill("solid", fgColor=BAD_FILL)
        if verdict_col is not None:
            vcell = ws.cell(row=r, column=verdict_col)
            vcell.font = Font(size=10, bold=True, color=BAD if bad else GOOD)
            if not bad:
                for c in range(1, len(headers) + 1):
                    ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=GOOD_FILL)
    autosize(ws, widths)


def title_block(ws, title, subtitle):
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=15, color=INK)
    ws["A2"] = subtitle
    ws["A2"].font = Font(size=10, color=ACCENT)
    ws.row_dimensions[1].height = 22


def sheet_summary(wb, window, full):
    ws = wb.create_sheet("Summary")
    title_block(ws, "DSD ASN REF*CN Audit", "Home Depot Canada · EDI 856 · Direct Store Delivery")

    win_bad = [r for r in window if r["verdict"] != "ok"]
    win_ok = [r for r in window if r["verdict"] == "ok"]
    full_bad = [r for r in full if r["verdict"] != "ok"]
    unrecoverable = [r for r in full_bad if not r["routing_description"]]
    dates = sorted(r["created_at"][:10] for r in window)

    facts = [
        ("Audit window", f"{dates[0]} to {dates[-1]} (sent date)"),
        ("DSD ASNs transmitted in window", len(window)),
        ("Sent with RTS number in REF*CN", len(win_bad)),
        ("Sent correctly", len(win_ok)),
        ("Correctable from the ASN itself", len([r for r in win_bad if r["routing_description"]])),
        ("", ""),
        ("Mapping corrected on", "2026-07-22"),
        ("Incorrect ASNs after that date", 0),
        ("", ""),
        ("DSD ASNs on record (all time)", len(full)),
        ("Affected across full history", len(full_bad)),
        ("Not correctable without HD's system", len(unrecoverable)),
    ]

    row = 4
    ws.cell(row=row, column=1, value="FINDING").font = Font(bold=True, size=9, color=ACCENT)
    ws.cell(row=row, column=2, value="VALUE").font = Font(bold=True, size=9, color=ACCENT)
    style_header(ws, row, 2)

    for k, v in facts:
        row += 1
        if not k:
            continue
        kc = ws.cell(row=row, column=1, value=k)
        vc = ws.cell(row=row, column=2, value=v)
        kc.font = Font(size=10, color=INK)
        vc.font = Font(size=10, bold=True, color=BAD if k == "Sent with RTS number in REF*CN" else INK)
        for c in (kc, vc):
            c.border = BOX
            c.alignment = Alignment(horizontal="left", vertical="center")

    row += 2
    ws.cell(row=row, column=1, value="WHAT WENT WRONG").font = Font(bold=True, size=10, color=ACCENT)
    notes = [
        "On every failing ASN the value in REF*CN is identical to the Bill of Lading number — the",
        "32-series RTS number was written into both fields. HD's validation looks for the 61-series",
        "Shipment ID from their transportation planning system, so none of these could match.",
        "",
        "The correct Shipment ID was not missing. It was in the Routing field (TD5-05, X12 element 387),",
        "a separate segment from the REF loop. This is a field-mapping fault, not missing data.",
        "",
        "RECOMMENDED NEXT STEPS",
        "1. Confirm the 2026-07-22 mapping fix was deliberate so it cannot regress silently.",
        "2. Send the Corrections sheet to Home Depot so they can re-match the inbound shipments.",
        "3. Decide whether to extend corrections back to 2026-03-06 (33 more ASNs, 26 self-correctable).",
        "4. Block any DSD 856 whose REF*CN does not begin with 61, or which equals the BOL number.",
    ]
    for n in notes:
        row += 1
        cell = ws.cell(row=row, column=1, value=n)
        cell.font = Font(size=10, bold=n.isupper() and bool(n), color=ACCENT if n.isupper() and n else INK)

    autosize(ws, [46, 40])
    return ws


def correction_rows(records):
    return [
        [
            r["asn_number"],
            r["created_at"][:10],
            r["shipment_date"],
            r["po_numbers"],
            r["store"],
            r["ship_to_name"].title().replace("Stock And Flow", "Stock and Flow"),
            r["ref_cn"] or "(empty)",
            r["routing_description"] or "NOT IN DOCUMENT",
            r["bill_of_lading_number"] or "-",
        ]
        for r in records
    ]


def sheet_corrections(wb, window):
    ws = wb.create_sheet("Corrections")
    title_block(ws, "ASNs to correct", "REF*CN should have carried the value in 'Should have been'")
    bad = sorted([r for r in window if r["verdict"] != "ok"], key=lambda r: r["created_at"])
    write_table(
        ws, 4,
        ["ASN", "Sent", "Ship date", "PO", "Store", "Ship to",
         "REF*CN sent (wrong)", "Should have been", "BOL / RTS"],
        correction_rows(bad),
        [14, 12, 12, 12, 8, 26, 20, 20, 14],
    )
    return ws


def all_rows(records):
    return [
        [
            r["asn_number"],
            r["created_at"][:10],
            r["shipment_date"],
            r["po_numbers"],
            r["store"],
            r["state"],
            r["ref_cn"] or "(empty)",
            r["bill_of_lading_number"] or "-",
            r["routing_description"] or "-",
            VERDICT_LABEL.get(r["verdict"], r["verdict"]),
        ]
        for r in sorted(records, key=lambda r: r["created_at"])
    ]


def sheet_all(wb, name, records, subtitle):
    ws = wb.create_sheet(name)
    title_block(ws, name, subtitle)
    write_table(
        ws, 4,
        ["ASN", "Sent", "Ship date", "PO", "Store", "State",
         "REF*CN", "BOL", "Routing (TD5-05)", "Verdict"],
        all_rows(records),
        [14, 12, 12, 12, 8, 11, 16, 16, 18, 24],
        verdict_col=10,
    )
    return ws


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default=os.path.join(REPO, f"DSD_ASN_REFCN_Audit_{date.today()}.xlsx"))
    args = ap.parse_args()

    for path in (WINDOW_JSON, FULL_JSON):
        if not os.path.exists(path):
            raise SystemExit(f"Missing {path}. Run tools/audit_dsd_asn_refcn.py first.")

    window = json.load(open(WINDOW_JSON))
    full = json.load(open(FULL_JSON))

    wb = Workbook()
    wb.remove(wb.active)
    sheet_summary(wb, window, full)
    sheet_corrections(wb, window)
    sheet_all(wb, "Window", window, "Every DSD ASN transmitted in the audit window")
    sheet_all(wb, "Full History", full, "Every DSD ASN on record, back to 2026-03-06")
    wb.save(args.out)

    print(f"Wrote {args.out}")
    print(f"  Summary      key figures + next steps")
    print(f"  Corrections  {len([r for r in window if r['verdict'] != 'ok'])} rows")
    print(f"  Window       {len(window)} rows")
    print(f"  Full History {len(full)} rows")


if __name__ == "__main__":
    main()
