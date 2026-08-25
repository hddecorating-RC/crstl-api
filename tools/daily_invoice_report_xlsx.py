"""Daily invoice report for accounting, as a .xlsx workbook.

One row per invoice -- invoice, type, province, date, subtotal, discounts,
charges, tax, total -- plus a total row. The Summary sheet carries the same
money split by type and lists anything whose parts do not add up to the total
HD stated, because an invoice that does not balance is worth seeing before it
is keyed rather than after.

Reads the raw 810 payload from Crstl rather than going through the app, so it
runs regardless of which version is deployed. Accepted only: Drafts are the
warehouse's superseded resubmissions and would duplicate rows.

Tax reaches us by two routes depending on the flavor, and both are read:

    Dropship / Wholesale   tax_information (TXI)   CG=GST  ST=QST  VA=HST
    DSD                    SAC codes               D360=GST  H680=QST
                                                   H770=HST  H850=ECO

That SAC reading follows the HD revision confirmed with Crstl in Aug 2026 and
HD's own rejection of INV40852200, not the v6.10 PDF in this repo, which
predates H770 and routes HST to H680. See app/sac_codes.py on the tax branch.

Dropship invoices carry no ship-to, so the province is recovered from the 850
PO; DSD carries its own.

Usage:
    python3 tools/daily_invoice_report_xlsx.py                  # yesterday
    python3 tools/daily_invoice_report_xlsx.py --date 2026-08-24
    python3 tools/daily_invoice_report_xlsx.py --from 2026-08-01 --to 2026-08-24
    python3 tools/daily_invoice_report_xlsx.py --out /path/to/file.xlsx
"""
import argparse
import os
import pathlib
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date, datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, REPO)

from app.crstl import CrstlClient          # noqa: E402
from app.main import load_env              # noqa: E402

num = CrstlClient._parse_float             # strips thousands separators

# Tax vocabularies. Crstl's own UI spells CG out as "Federal Value Added Tax
# GST On Goods"; these readings come from Crstl and from HD, not from matching
# amounts against province rates.
TXI_KIND = {"CG": "GST", "ST": "QST", "VA": "HST"}
SAC_KIND = {"D360": "GST", "H680": "QST", "H770": "HST",
            "H850": "ECO", "F240": "ECO", "G090": "ECO", "G100": "ECO"}

# Matches app/main.py's rate-check convention: a 2-cent floor, widening to 0.2%
# on larger invoices. HD rounds each line to the cent, so summing lines can
# drift a cent or two from the stated total without anything being wrong.
def tolerance_for(amount):
    return max(0.02, round(0.002 * abs(amount), 2))


MONEY = "#,##0.00"
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
BAD_FILL = PatternFill("solid", fgColor="FFC7CE")


def flavor_of(meta):
    """Crstl sends three flavors. Wholesale is rare but real, and collapsing it
    into Dropship would mislabel the column accounting reads to tell the
    channels apart."""
    raw = str(meta.get("trading_partner_flavor") or "")
    if "DSD" in raw or "Direct Store" in raw:
        return "DSD"
    if "Wholesale" in raw:
        return "Wholesale"
    if "Dropship" in raw:
        return "Dropship"
    return raw or "Unknown"


def extract(detail, po_index):
    """Flatten one 810 payload into the row accounting needs."""
    meta = detail.get("metadata", {}) or {}
    edi = (detail.get("file", {}) or {}).get("generic_json_edi", {}) or {}
    heading = edi.get("heading", {}) or {}
    summary = edi.get("summary", {}) or {}

    po = str(meta.get("source_document_reference_id") or heading.get("purchase_order_number") or "")
    ship_to = heading.get("ship_to") or {}
    province = (str(ship_to.get("state_province") or "").upper()
                or str((po_index.get(po) or {}).get("province") or "").upper())

    subtotal = round(sum(
        num((e.get("baseline_item_data_invoice", {}) or {}).get("quantity_invoiced"))
        * num((e.get("baseline_item_data_invoice", {}) or {}).get("unit_price"))
        for e in (edi.get("detail", {}) or {}).get("baseline_item_data_invoice_loop", []) or []
    ), 2)

    deductions, charges, taxes, unknown = {}, {}, {}, {}
    for entry in summary.get("service_promotion_allowance_or_charge_information_loop", []) or []:
        sac = entry.get("service_promotion_allowance_or_charge_information", {}) or {}
        code = sac.get("service_promotion_allowance_or_charge_code") or "(no code)"
        amt = num(sac.get("amount"))
        indicator = sac.get("allowance_or_charge_indicator")
        # The indicator decides, not the code. HD has sent D360 -- a tax code --
        # with indicator "A" on 3 invoices, and reading it as tax puts those
        # 70.52 on the wrong side: INV40831749 reconciles at 5,317.43 only if
        # the allowance is honoured as an allowance. Report what was sent.
        if indicator == "A":
            deductions[code] = round(deductions.get(code, 0) + amt, 2)
        elif indicator == "C":
            kind = SAC_KIND.get(code)
            if kind:
                taxes[kind] = round(taxes.get(kind, 0) + amt, 2)
            else:
                charges[code] = round(charges.get(code, 0) + amt, 2)
        else:
            # Neither allowance nor charge. Its sign is unknowable, so it is
            # shown but kept out of the reconciliation rather than guessed at.
            unknown[code] = round(unknown.get(code, 0) + amt, 2)

    for entry in summary.get("tax_information", []) or []:
        kind = TXI_KIND.get(entry.get("tax_type_code")) or entry.get("tax_type_code") or "(no code)"
        taxes[kind] = round(taxes.get(kind, 0) + num(entry.get("monetary_amount")), 2)

    # metadata.value is Crstl's canonical total (app/crstl.py says so); fall
    # back to the EDI summary only when it is genuinely absent, not when it is
    # a legitimate 0.00.
    stated = num(meta.get("value") if meta.get("value") is not None
                 else (summary.get("total_monetary_value_summary", {}) or {}).get("total_amount"))

    # An 810 with no parsable line loop reports subtotal 0; the app treats the
    # stated total as the subtotal in that case rather than under-reporting.
    if subtotal == 0.0:
        subtotal = stated

    computed = round(subtotal - sum(deductions.values()) + sum(charges.values())
                     + sum(taxes.values()), 2)

    return {
        "invoice": str(meta.get("reference_id") or heading.get("invoice_number") or ""),
        "date": heading.get("invoice_date") or "",
        "po": po,
        "flavor": flavor_of(meta),
        "province": province,
        "subtotal": subtotal,
        "deductions": deductions,
        "charges": charges,
        "taxes": taxes,
        "unknown": unknown,
        "stated": stated,
        "computed": computed,
        "variance": round(stated - computed, 2),
    }


def reconciles(row):
    return abs(row["variance"]) <= tolerance_for(row["stated"])


def build_workbook(rows, out_path, window):
    """Columns are driven by what the invoices actually carry.

    Every SAC code and every tax code present in the window gets its own
    column, so nothing on an invoice is invisible because a column was not
    anticipated. Freight and fee charges have never been transmitted to date,
    but if one arrives it appears as its own column rather than disappearing
    into a total that then fails to add up.
    """
    ded_codes = sorted({c for r in rows for c in r["deductions"]})
    chg_codes = sorted({c for r in rows for c in r["charges"]})
    tax_kinds = sorted({k for r in rows for k in r["taxes"]})
    unk_codes = sorted({c for r in rows for c in r["unknown"]})

    def tot(rs, key):
        return round(sum(sum(r[key].values()) for r in rs), 2)

    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"

    headers = (["Invoice", "Type", "Province", "Invoice Date", "Subtotal",
                "Discounts"] + [f"Ded {c}" for c in ded_codes]
               + ["Charges"] + [f"Chg {c}" for c in chg_codes]
               + ["Tax"] + [f"Tax {k}" for k in tax_kinds]
               + (["Unclassified"] + [f"Unc {c}" for c in unk_codes] if unk_codes else [])
               + ["Total"])
    ws.append(headers)
    for cell in ws[1]:
        cell.fill, cell.font = HDR_FILL, HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r in rows:
        row = ([r["invoice"], r["flavor"], r["province"], r["date"], r["subtotal"],
                -round(sum(r["deductions"].values()), 2) or 0]
               + [-r["deductions"][c] if c in r["deductions"] else None for c in ded_codes]
               + [round(sum(r["charges"].values()), 2) or 0]
               + [r["charges"].get(c) for c in chg_codes]
               + [round(sum(r["taxes"].values()), 2) or 0]
               + [r["taxes"].get(k) for k in tax_kinds])
        if unk_codes:
            row += [round(sum(r["unknown"].values()), 2) or 0] + [r["unknown"].get(c) for c in unk_codes]
        row.append(r["stated"])
        ws.append(row)

    last = ws.max_row + 1
    total_row = (["Total", "", "", "", round(sum(r["subtotal"] for r in rows), 2),
                  -tot(rows, "deductions")]
                 + [-round(sum(r["deductions"].get(c, 0) for r in rows), 2) or None for c in ded_codes]
                 + [tot(rows, "charges")]
                 + [round(sum(r["charges"].get(c, 0) for r in rows), 2) or None for c in chg_codes]
                 + [tot(rows, "taxes")]
                 + [round(sum(r["taxes"].get(k, 0) for r in rows), 2) or None for k in tax_kinds])
    if unk_codes:
        total_row += [tot(rows, "unknown")] + [round(sum(r["unknown"].get(c, 0) for r in rows), 2) or None
                                               for c in unk_codes]
    total_row.append(round(sum(r["stated"] for r in rows), 2))
    ws.append(total_row)
    for cell in ws[last]:
        cell.font = Font(bold=True)
        cell.border = Border(top=Side(style="thin"))

    for row in ws.iter_rows(min_row=2, min_col=5):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = MONEY
    widths = [18, 11, 10, 14, 13] + [12] * (len(headers) - 6) + [13]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "E2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last - 1}"

    s = wb.create_sheet("Summary")
    s.append(["Invoiced", window]); s["A1"].font = Font(bold=True, size=13)
    s.append(["Source", "Crstl 810 transactions, state = Accepted"])
    s.append([])
    s.append(["Type", "Invoices", "Subtotal", "Discounts", "Charges", "Tax", "Total"])
    for cell in s[4]:
        cell.fill, cell.font = HDR_FILL, HDR_FONT
    for flavor in sorted({r["flavor"] for r in rows}):
        sub = [r for r in rows if r["flavor"] == flavor]
        s.append([flavor, len(sub), round(sum(r["subtotal"] for r in sub), 2),
                  -tot(sub, "deductions"), tot(sub, "charges"), tot(sub, "taxes"),
                  round(sum(r["stated"] for r in sub), 2)])
    s.append(["Total", len(rows), round(sum(r["subtotal"] for r in rows), 2),
              -tot(rows, "deductions"), tot(rows, "charges"), tot(rows, "taxes"),
              round(sum(r["stated"] for r in rows), 2)])
    for cell in s[s.max_row]:
        cell.font = Font(bold=True)

    unbalanced = [r for r in rows if not reconciles(r)]
    s.append([])
    s.append(["Do not match HD's total", len(unbalanced)])
    s[s.max_row][0].font = Font(bold=True)
    if unbalanced:
        s[s.max_row][1].fill = BAD_FILL
        s.append(["Invoice", "HD total", "Adds up to", "Difference"])
        for cell in s[s.max_row]:
            cell.font = Font(bold=True)
        for r in unbalanced:
            s.append([r["invoice"], r["stated"], r["computed"], r["variance"]])
    for row in s.iter_rows(min_row=5):
        for cell in row:
            if isinstance(cell.value, (int, float)) and cell.column > 2:
                cell.number_format = MONEY
    for i, w in enumerate((24, 14, 14, 14, 12, 12, 14), start=1):
        s.column_dimensions[get_column_letter(i)].width = w

    wb.save(out_path)
    return out_path


def _fetch_details(client, transactions):
    """Fetch details in parallel, skipping records that fail rather than losing
    the whole report to one bad one -- the same trade-off fetch_invoices makes."""
    out = []
    with ThreadPoolExecutor(max_workers=CrstlClient.MAX_WORKERS) as pool:
        futures = {}
        for t in transactions:
            tid = t.get("id") or (t.get("metadata") or {}).get("id")
            if not tid:
                print(f"WARNING: skipping transaction with no id: {t.get('reference_id')}")
                continue
            futures[pool.submit(client._fetch_transaction_detail, tid)] = t
        for fut in as_completed(futures):
            try:
                out.append(fut.result())
            except Exception as exc:
                print(f"WARNING: could not fetch {futures[fut].get('reference_id')}: {exc}")
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--date", help="single invoice_date (YYYY-MM-DD); defaults to yesterday")
    ap.add_argument("--from", dest="date_from", help="start of an inclusive range")
    ap.add_argument("--to", dest="date_to", help="end of an inclusive range")
    ap.add_argument("--out", help="output .xlsx path")
    args = ap.parse_args()

    if args.date_from or args.date_to:
        lo = args.date_from or "0000-00-00"
        hi = args.date_to or "9999-99-99"
    else:
        day = args.date or (date.today() - timedelta(days=1)).isoformat()
        lo = hi = day
    window = lo if lo == hi else f"{lo} to {hi}"

    load_env(os.path.join(REPO, ".env"))
    client = CrstlClient(base_url=os.environ.get("CRSTL_BASE_URL", "https://api.crstl.so/v2"),
                         api_key=os.environ["CRSTL_API_KEY"])

    print(f"Fetching 810 transactions for {window} ...")
    listing = client._fetch_all_transactions("810")

    accepted, unknown_states = [], {}
    for t in listing:
        state = (t.get("state") or {}).get("value")
        if state == "Accepted":
            accepted.append(t)
        elif state != "Draft":
            unknown_states[state] = unknown_states.get(state, 0) + 1
    if unknown_states:
        # Same reasoning as app/main.py's _reportable: an unrecognised state
        # must not vanish silently, or a new "good" state would quietly stop
        # reaching accounting.
        print(f"WARNING: withheld invoices with unrecognised status {unknown_states} — "
              f"if one of these is reportable, this filter needs updating")

    # Narrow by created_at before fetching details. A generous buffer keeps
    # invoices whose invoice_date trails their creation; the exact filter on
    # invoice_date still happens after extraction.
    def created(t):
        return str(t.get("created_at") or "")[:10]
    buf_lo = (datetime.fromisoformat(lo).date() - timedelta(days=14)).isoformat() if lo[0].isdigit() and lo != "0000-00-00" else lo
    buf_hi = (datetime.fromisoformat(hi).date() + timedelta(days=14)).isoformat() if hi[0].isdigit() and hi != "9999-99-99" else hi
    candidates = [t for t in accepted if not created(t) or buf_lo <= created(t) <= buf_hi]

    details = _fetch_details(client, candidates)

    # Only look up the POs actually needed; fetch_po_provinces() crawls every
    # 850 on record, which is wasted work for a one-day report.
    need_po = {str((d.get("metadata") or {}).get("source_document_reference_id") or "")
               for d in details}
    po_index = {}
    pos = [p for p in client._fetch_all_transactions("850")
           if str(p.get("reference_id") or "") in need_po]
    for detail in _fetch_details(client, pos):
        st = ((detail.get("file", {}) or {}).get("generic_json_edi", {})
              .get("heading", {}) or {}).get("ship_to", {}) or {}
        ref = str((detail.get("metadata") or {}).get("reference_id") or "")
        if ref and st.get("state_province"):
            po_index[ref] = {"province": str(st["state_province"]).upper()}

    rows, undated = [], 0
    for detail in details:
        row = extract(detail, po_index)
        if not row["date"]:
            undated += 1
            print(f"WARNING: {row['invoice']} has no invoice_date and is not in any window")
            continue
        if lo <= row["date"] <= hi:
            rows.append(row)
    rows.sort(key=lambda r: (r["flavor"], r["invoice"]))

    if not rows:
        # A quiet day is not a failure. A cron wrapper must not read it as one.
        print(f"No Accepted invoices dated {window}.")
        return 0

    out = args.out or os.path.join(REPO, ".tmp",
                                   f"HD_Invoices_{lo}.xlsx" if lo == hi
                                   else f"HD_Invoices_{lo}_to_{hi}.xlsx")
    pathlib.Path(out).parent.mkdir(parents=True, exist_ok=True)
    build_workbook(rows, out, window)

    bad = sum(1 for r in rows if not reconciles(r))
    by_flavor = ", ".join(f"{sum(1 for r in rows if r['flavor'] == f)} {f}"
                          for f in sorted({r["flavor"] for r in rows}))
    print(f"{len(rows)} invoices ({by_flavor})")
    print(f"{bad} do not reconcile" if bad else "all invoices reconcile")
    print(f"-> {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
